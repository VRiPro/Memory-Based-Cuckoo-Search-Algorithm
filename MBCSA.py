'''__________ALGORITHME PROJET CL02__________'''

# Importation des bibliothèques
import numpy as np
from openpyxl import load_workbook
import random
from tqdm import tqdm
import pandas as pd
import time
import tkinter as tk

def lecteur():
    '''Procédure appelée au début du programme
    Lit les données depuis un des documents Excel (selon le problème souhaité)'''
    
    #Lecture des données dans des variables globales :
    global wb; wb = load_workbook("Données{}.xlsx".format(pb)) # Fichier Excel
    global ws; ws = wb["Feuil1"] #Page 1 du fichier Excel
    global ws2; ws2 = wb["Feuil2"] #Page 2 du fichier Excel
    global T; T = ws.cell(1,2).value # Nombre de tâches
    global W; W = ws.cell(2,2).value # Nombre de stations
    global R; R = ws.cell(3,2).value # Nombre de robots  
    global es; es = np.zeros(R,dtype=float) # Énergie en standby
    for r in range(R):
        es[r] = ws.cell(r+2,5).value  
        
    global t; t = np.zeros([R,T],dtype = float) # Durée de traitement
    global e; e = np.zeros([R,T],dtype = float) # Énergie de traitement
    for r in range(R):
        for i in range(T):
            t[r,i] =  ws.cell(i+3,r+8).value
            e[r,i] =  ws.cell(i+3,r+8+R).value
    
    global P; P = np.zeros([T,T],dtype=bool) # Matrice de précédence
    for i in range(T):
        for j in range(T):
            P[i,j] = ws2.cell(i+3,j+2).value
            
def initialisation():    
    '''Procédure utilisée au début de l'algorithme :
    Remplit une variable globale X avec une population initiale de N nids/oeufs réalisables'''
    global X; X = [] # Population de N nids
    global M; M = [] # Mémoire
    
    for n in range(N):
        solution = hasard_faisable()
        X.append(list(solution))       
 
def hasard_faisable():
    '''Fonction appelée lors de l'initialisation
    Renvoie un ordonnancement faisable selon les relations de précédence'''
    tache_visitee = np.zeros(T,dtype=bool)
    ordonnancement = []
    for ordre in range(T):
        liste_taches_possibles =[]
        for tache in range(T):
            if not tache_visitee[tache]:
                tache_possible = True
                for predecesseur in range(T):
                    if P[tache,predecesseur]==1 and not tache_visitee[predecesseur]:
                        tache_possible=False
                if tache_possible:
                    liste_taches_possibles.append(tache+1)
        indice_aleatoire = random.randint(0,len(liste_taches_possibles)-1)
        ordonnancement.append(liste_taches_possibles[indice_aleatoire])
        tache_visitee[liste_taches_possibles[indice_aleatoire]-1] = True
    return ordonnancement

def affectations(x,NRJ,COUNT):
    ''' Fonction traduisant les résultats de la fonction best_robot (NRJ et COUNT)
    en des listes claires concernant les affectations des robots et tâches aux stations'''
    robots = np.zeros(W,dtype=int)
    nb_taches = np.zeros(W,dtype=int)
    for station in range(W):
        nb_taches[station] = np.amax(COUNT[station])
        energie_min = 100000
        for robot in range(R):
            if COUNT[station][robot] == nb_taches[station] and NRJ[station][robot]<energie_min:
                energie_min = NRJ[station][robot]
                robots[station]=robot+1
    return robots,nb_taches

def fitness(x):
    '''Fonction qui renvoie le fitness (=consommation d'énergie totale) d'un ordonnancement x donné
    Pour cela il faut également avoir l'affectation des tâches aux stations
    et l'affectation des robots aux stations
    '''
    NRJ,COUNT = best_robot(x)
    robots,nb_taches = affectations(x,NRJ,COUNT)
    fit_process = 0
    fit_standby = 0
    for robot in range(W):
        for i in x[sum(nb_taches[:robot]):sum(nb_taches[:robot+1])]:
            fit_process+=e[robots[robot]-1][i-1]
        fit_standby += Cycle_Time(x,robots,nb_taches)[robot]*es[robots[robot]-1]
    return fit_process+fit_standby          

def Cycle_Time(x,robots,nb_taches):
    ''' Calcul de la différence entre le CT max et celui d'une station
    Pour le calcul de l'energie en standby dans la fonction fitness'''
    CT=np.zeros(W)
    for robot in range(W):
        for i in x[sum(nb_taches[:robot]):sum(nb_taches[:robot+1])]:
            CT[robot]=CT[robot]+t[robots[robot]-1][i-1]
    return np.max(CT)-CT

def Cycle_Time_Max(x,robots,nb_taches):
    ''' Fonction pour afficher à la fin le CycleTime max'''
    CT=np.zeros(W)
    for robot in range(W):
        for i in x[sum(nb_taches[:robot]):sum(nb_taches[:robot+1])]:
            CT[robot]=CT[robot]+t[robots[robot]-1][i-1]
    return np.max(CT)

def reparation(x):
    ''' Fonction qui recoit un ordonnancement non faisable et en renvoie un faisable'''
    verif_pred = tester_faisable(x)
    while verif_pred != True :
        x[x.index(verif_pred[1])]=verif_pred[0]#obligation d'avoir index(verif_pred[1]) changé avant car verif_pred[1] doit etre avant verif_pred[0] or index(verif_pred[1]) choisit l'index du premier verif_pred[1] qu'il trouve.
        x[x.index(verif_pred[0])]=verif_pred[1]
        verif_pred = tester_faisable(x)
    return x

def tester_faisable(ordonnancement):
    '''Fonction utilisée par le mécanisme de réparation
    Teste si un ordonnancement donné est faisable selon les relations de précédence
    Si non, renvoie les indices des deux premières tâches à échanger trouvées'''
    
    resultat = True
    tache_visitee = np.zeros(T,dtype=bool)

    for tache in reversed(ordonnancement):
        tache_actuelle = tache-1
        tache_visitee[tache_actuelle] = True
        for tache2 in range(T):
            if P[tache_actuelle,tache2] and tache_visitee[tache2]:
                resultat = False
                return tache_actuelle+1, tache2+1 # tache2+1 est le prédécesseur de tache_actuelle+1
            # ces deux tâches sont renvoyées à la fonction réparation pour être échangées
    return resultat 

def generation_voisinage(x):
    '''Fonction qui génère un voisinage de nb_voisins voisins d'une solution (ordonnancement) donnée
    Les voisins correspondent à une mutation swap suivie d'un mécanisme de réparation
    Tous les voisins sont uniques'''
    voisinage = []
    i = 0
    while i < nb_voisins:
        voisin = x[:]
        
        indice_tache1 = random.randint(0,T-1)
        indice_tache2 = random.randint(0,T-1)
        while indice_tache2 == indice_tache1:
            indice_tache2 = random.randint(0,T-1)
        tache1 = voisin[indice_tache1]
        tache2 = voisin[indice_tache2]
        voisin[indice_tache1] = tache2
        voisin[indice_tache2] = tache1
        
        voisin = reparation(voisin)
        
        if not voisin in voisinage:
            voisinage.append(voisin)
            i+=1
            
    return voisinage
            
def cross_over_ordo(p1,p2):
    '''Fonction qui renvoie les 2 enfants d'un cross-over entre deux parents donnés (des ordonnancements)'''
    alea = np.random.randint(1,T//2+1)
    verif_enf1=np.zeros(T,bool)
    verif_enf2=np.zeros(T,bool)
    enf1=[]
    enf2=[]
    for i in range(alea):
        enf1.append(p1[i])
#enf = 1er partie de parent 1 puis on complete enf1 par les taches manquante dans l'ordre d'apparition de parent 2
        enf2.append(p2[i])
        verif_enf1[p1[i]-1]=True
        verif_enf2[p2[i]-1]=True
    #enfant 1 egale au debut de parent 1 et se complet par le reste de parent 2
    for i in range(T):
        if verif_enf2[p1[i]-1]==False:
            enf2.append(p1[i])
        if verif_enf1[p2[i]-1]==False:
            enf1.append(p2[i])
    
    #Mécanisme de réparation :
    enf1 = reparation(enf1)
    enf2 = reparation(enf2)
    
    return enf1,enf2

def selection():
    ''' Sélection : une proportion PA des oeufs (solutions) est abandonnée et remplacée'''
    global max_search
    couples=[]
    X.sort(key=fitness)
    
    NA = int(PA*N) # Nombre de nids abandonnés
    count=NA
    while count !=0:#tq l'on pas pas PA enfants on en génére
        if len(couples)==(N-NA)*(N-NA-1):
            return
        while len(couples)<(N-NA)*(N-NA-1) and count!=0:#permet de d'arreter le systeme si tout les couples possibles sont générés        
            alea1,alea2=np.random.randint(0,N-NA),np.random.randint(0,N-NA)
            if alea1!=alea2:
                if (alea1,alea2) not in couples:#on regarde si on a deja fait ce croisement
                    couples.append((alea1,alea2))
                    couples.append((alea2,alea1))
                    Cr_Ov=cross_over_ordo(X[alea1],X[alea2])#on enregistre le cross over car celui ci peut comporter une part d'aleatoire
                    for i in range(2):#On vérifie 1 par un si les enfants de notre croisement ne sont pas deje en mémoire et on les rajoutes à la population si nécessaire
                        if count>0:
                            if verif_memo(Cr_Ov[i])==False:
                                max_search=max_search_init
                                count-=1
                                X[-count-1]=Cr_Ov[i]
                            else:
                                max_search-=1
                            if max_search==0:
                                return
    #if count >0:
        #print("impossible de remplacer une partie de notre population")
    return 

def verif_memo(x):
    '''Fonction pour vérifier si une solution donnée se trouve déjà dans la mémoire'''
    if x in M:
        return True#dans la memoire
    else:
        return False#pas dans la mémoire

def put_memo(x):
    M.append(x)


def best_robot(x):
    ''' Fonction déterminant des bonnes affectations tâches/station et robot/station
    à partir d'un ordonnancement des tâches x donné
    Il provient d'un autre article auquel font référence les auteurs'''
    E0=np.sum(np.min(e.T,axis=1))/W
    NRJ=np.zeros((W,R),dtype=float)
    COUNT=np.zeros((W,R),dtype=int)
    ok = False
    while ok != True:
        for w in range(W):
            if w!=0:
                startw=sum(np.max(COUNT[i])for i in range(w))
            else:
                startw=0
            for r in range(R):
                count=0
                E=0
                while E<=E0 and startw+count<T:
                    E+=e[r][x[startw+count]-1]
                    count+=1
                if startw+count<T:                    
                    NRJ[w][r]=E-e[r][x[startw+count-1]-1]
                    count -= 1
                    COUNT[w][r]=count
                else:
                    NRJ[w][r]=E
                    COUNT[w][r]=count
            if np.max(np.sum(COUNT,axis=0))==T:
                return NRJ,COUNT
                
        if count==T:
            ok=True
        else:
            E0+=E0step              #choisir me meilleur ordonancement
    return NRJ,COUNT


#_____MBCSA____________________________________________________________________________
def main_algo():
    '''Algorithme principal de la heuristique MBCSA tel que présenté dans l'article''' 
    global max_search,max_generations
    max_search = max_search_init
    max_generations=max_generations_get
    initialisation()
    Best_x = X[0]
    while max_generations >=0:
        while max_search >=0:
            alea = np.random.randint(0,N-1)
            x=X[alea]
            if not x in M:
                max_search=max_search_init
                break
            else :
                max_search-=1
                if max_search==0:
                    return Best_x
        V=generation_voisinage(x)
        M.append(V)
        M.append(x)
        fitV=[fitness(v) for v in V]
        alea = np.random.randint(0,N-1)
        h=X[alea]
        C=V[np.argmin(fitV)]
        if fitness(h)>fitness(C):
            X[alea]=C
        selection()
        X.sort(key=fitness)
        if fitness(X[0])<fitness(Best_x):
            Best_x=X[0]
        max_generations-=1
    return Best_x
#______________________________________________________________________________________

#_____MAIN_____________________________________________________________________________
def main():
    '''
    Procédure principale qui récupère les données, exécute plusieurs fois l'algorithme 
    et exporte les résultats dans un document Excel
    '''
    print('Problème ',pb)
    lecteur();print("Données récupérées")
    print("*** Lancement des algorithmes ***")
    resultat_total =[]
    for algo in tqdm(range(nbr_exe)):
        ###########################
        # EXÉCUTION :
        begin = time.time()
        best_ordo = main_algo()
        end=time.time()
        elapsed=end-begin
        ###########################
        NRJ,COUNT = best_robot(best_ordo)
        robots,nb_taches = affectations(best_ordo, NRJ, COUNT)
        CT = Cycle_Time_Max(best_ordo, robots, nb_taches)
        
        resultat_algo = [robots,nb_taches,best_ordo,fitness(best_ordo),CT,elapsed]

        resultat_total.append(resultat_algo)
    tab_final = pd.DataFrame(data=resultat_total,columns=['Robots','NbTaches','Ordonnancement','Fitness','CycleTime','TempsAlgo (s)'])
    tab_final.to_excel('Résultat.xlsx')
    print('*** Terminé ***')
    print('Le résultat se trouve dans le fichier Excel Résultat.xlsx')
    resul = tk.Tk()
    resul.geometry("350x80")
    resul.title('RESULTATS')
    tk.Label(resul, text="LES RESULTATS SONT DANS LE TABLEUR 'Resultats.xlsx' ! ").pack(fill='x')
    boutonresul=tk.Button(resul,text="Quitter",command=resul.destroy,bg="red",cursor="X_cursor").pack(expand=1,fill="both")
#______________________________________________________________________________________    


# root window
root = tk.Tk()
#root.geometry("300x333")
root.geometry("300x500")
root.resizable(False, False)
root.title('PARAMETRES')

tk.Label(root, text="N:").pack(fill='x')
n_txt=tk.Entry(root)
n_txt.insert(-1,20)
n_txt.pack()
 #Nombre de nids hôtes

tk.Label(root, text="max_search:").pack(fill='x')
ms_txt=tk.Entry(root)
ms_txt.insert(-1,500)
ms_txt.pack()

tk.Label(root, text="PA:").pack()
PA_txt=tk.Entry(root)
PA_txt.insert(-1,0.15)
PA_txt.pack()

tk.Label(root, text="nb_voisins:").pack(fill='x')
nbv_txt=tk.Entry(root)
nbv_txt.insert(-1,10)
nbv_txt.pack()

tk.Label(root, text="max_generation:").pack(fill='x')
mx_gen=tk.Entry(root)
mx_gen.insert(-1,200)
mx_gen.pack()

tk.Label(root, text="E0 step:").pack(fill='x')
E0_txt=tk.Entry(root)
E0_txt.insert(-1,0.05)
E0_txt.pack() # Numéro du problème à traiter


tk.Label(root, text="Numero du Probleme:").pack(fill='x')
PB_txt=tk.Entry(root)
PB_txt.insert(-1,1)
PB_txt.pack()
 # Numéro du problème à traiter

tk.Label(root, text="Nombre d'executions:").pack(fill='x')
exe_txt=tk.Entry(root)
exe_txt.insert(-1,10)
exe_txt.pack()

def parametrage():
    ''' Première procédure à la suite du remplissage des paramètres'''
    global N,nb_voisins,max_search_init,PA,E0step,pb,nbr_exe,max_generations_get
    N =n_txt.get()
    N=int(N)
    nb_voisins = nbv_txt.get()
    nb_voisins=int(nb_voisins)
    max_search_init = ms_txt.get()
    max_search_init = int(max_search_init)
    PA = PA_txt.get()
    PA= float(PA)
    E0step = E0_txt.get()
    E0step = float(E0step)
    pb = PB_txt.get()
    pb = int(pb)
    nbr_exe= exe_txt.get()
    nbr_exe=int(nbr_exe)
    max_generations_get=mx_gen.get()
    max_generations_get=int(max_generations_get)
    root.destroy()
    ''' Lancement des exécutions :'''
    main()
    

bouton=tk.Button(root,text="Lancer l'algorithme",bg="green",command=parametrage,cursor="hand2")
tk.Label(root, text="").pack()
bouton.pack(expand=1,fill="x")


root.mainloop()